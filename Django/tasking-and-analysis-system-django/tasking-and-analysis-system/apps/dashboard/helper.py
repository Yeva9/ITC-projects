import datetime
import numpy as np
from django.contrib.auth.decorators import login_required
from django.contrib.auth.decorators import user_passes_test
from django.db import ProgrammingError
from django.http import HttpResponseNotFound
from django.shortcuts import render, HttpResponse
from django.template import loader
from openpyxl import Workbook

from apps.audit_trail.models import AuditTrail
from apps.authentication.models import User
from apps.constants import General, AuthenticationFields, DashboardFields, ReportsFields
from apps.reports.models import SuccessExport
from apps.tasks.models import Task
from apps.upload.models import Upload, UploadSuccess
from .manager import can_view_tab


def get_day_statistic(day, all_logins):
    """
        This function calculates users unique login count.
    """

    checked_users = []
    count = 0
    for info in all_logins:
        info_date = datetime.date(info.event_date.year, info.event_date.month, info.event_date.day)
        if info_date == day and info.user not in checked_users:
            checked_users.append(info.user)
            count += 1

    return count


def get_avg_readability(uploads):
    """
        This function returns the average readability of the uploaded data.
    """

    readability = []
    for it in uploads:
        readability.append(list(it.data.values()))

    readability = (np.mean(readability, axis=0, dtype=np.int)).tolist()  # average by column

    return readability


def get_day_readability(uploads):
    """
        This function returns days of uploaded data.
    """

    date = []
    for it in uploads:
        date = list(it.data.keys())  # get date
        break

    days_str = [i[0:2] for i in date]  # get days
    days = list(map(int, days_str))  # converting to int

    return days


def get_task_progress_count(request):
    """
    This function returns count of task progresses for the current users.
    """
    if request.user.role.name == 'Task creator':
        progress_count = {
            'count1': Task.objects.filter(progress=2, assigned_by_id=request.user.id).count(),
            'count2': Task.objects.filter(progress=4, assigned_by_id=request.user.id).count(),
            'count3': Task.objects.filter(progress=3, assigned_by_id=request.user.id).count(),
            'count4': Task.objects.filter(assigned_by_id=request.user.id).count(),
        }
    else:
        progress_count = {
            'count1': Task.objects.filter(progress=2, assigned_to_id=request.user.id).count() +
                        Task.objects.filter(progress=2, assigned_by_id=request.user.id).count(),
            'count2': Task.objects.filter(progress=4, assigned_to_id=request.user.id).count() +
                        Task.objects.filter(progress=4, assigned_by_id=request.user.id).count(),
            'count3': Task.objects.filter(progress=3, assigned_to_id=request.user.id).count() +
                        Task.objects.filter(progress=3, assigned_by_id=request.user.id).count(),
            'count4': Task.objects.filter(assigned_to_id=request.user.id).count() +
                        Task.objects.filter(assigned_by_id=request.user.id).count(),
        }

    return progress_count


def export_logs_to_xlsx(request):
    """
    This function downloads all logs as Excel file with a single worksheet
    """

    log_queryset = AuditTrail.objects.all().order_by('id')

    response = HttpResponse(
        content_type=ReportsFields.CONTENT_TYPE,
    )
    response['Content-Disposition'] = 'attachment; filename={date}-logs.xlsx'.format(
        date=datetime.datetime.now().strftime('%Y-%m-%d'),
    )
    workbook = Workbook()
    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = 'Event log'

    # Define the titles for columns
    columns = [
        'ID',
        'Event',
        'Event Description',
        'User',
        'Date',
    ]
    row_num = 1
    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
    # Iterate through all logs
    for log in log_queryset:
        row_num += 1
        # Define the data for each cell in the row
        row = [
            log.id,
            log.event_title,
            log.event_description,
            log.user,
            log.event_date,
        ]

        # Assign the data for each cell of the row
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(response)

    return response


def export_data_to_xlsx(request):
    """
        This function downloads uploaded data as Excel file with a single worksheet(by Task Creator)
    """

    data_queryset = Upload.objects.all().order_by('id')[:10]

    response = HttpResponse(
        content_type=ReportsFields.CONTENT_TYPE,
    )
    response['Content-Disposition'] = 'attachment; filename={date}-data.xlsx'.format(
        date=datetime.datetime.now().strftime('%Y-%m-%d'),
    )
    workbook = Workbook()
    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = 'Data'

    # Define the titles for columns
    columns = [
        'ID',
        'Substation',
        'Data',
    ]
    row_num = 1
    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
    # Iterate through all logs
    for dat in data_queryset:
        row_num += 1
        # Define the data for each cell in the row
        row = [
            dat.id,
            dat.farm,
            str(dat.data)
        ]

        # Assign the data for each cell of the row
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(response)

    return response
